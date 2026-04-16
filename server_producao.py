import sys, os, json, base64, tempfile, traceback, secrets, copy, shutil
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse
from datetime import datetime

# ══════════════════════════════════════
SENHA   = "acampamento2026"   # ← TROQUE
PORT    = int(os.environ.get("PORT", 8000))
DIR     = os.path.dirname(os.path.abspath(__file__))
MODELO  = os.path.join(DIR, 'VIII_MODELO.xlsx')
RESTRICOES_FILE = os.path.join(DIR, 'restricoes_tribos.json')
# ══════════════════════════════════════

sys.path.insert(0, DIR)

SESSOES = {}

def nova_sessao():
    t = secrets.token_hex(32); SESSOES[t] = True; return t

def sessao_valida(token):
    return bool(token and token in SESSOES)

# ── Carrega restrições ──
def load_restricoes():
    if os.path.exists(RESTRICOES_FILE):
        try:
            with open(RESTRICOES_FILE, encoding='utf-8') as f:
                return json.load(f)
        except: pass
    return {"restricoes": [], "obs": ""}

def save_restricoes(data):
    data['obs'] = f"Atualizado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    with open(RESTRICOES_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ── ZPL helpers ──
TRIBO_COR_NOME = {
    'Simeão':'AMARELA','Rubem':'VERMELHA','Judá':'LARANJA','Levi':'ROSA',
    'Benjamim':'AZUL','Issacar':'VERDE','Gad':'MARROM','Zabulom':'PRETA','EFRAIM':'CINZA',
}

def zpl_safe(s):
    r = {'ã':'a','â':'a','á':'a','à':'a','é':'e','ê':'e','è':'e','í':'i','î':'i',
         'ó':'o','ô':'o','õ':'o','ú':'u','û':'u','ç':'c',
         'Ã':'A','Â':'A','Á':'A','É':'E','Ê':'E','Í':'I','Ó':'O','Ô':'O','Õ':'O',
         'Ú':'U','Û':'U','Ç':'C'}
    for k,v in r.items(): s=s.replace(k,v)
    return s

def quebra_nome(nome, max_chars=18):
    if len(nome) <= max_chars: return nome, ''
    meio = max_chars
    while meio > 0 and nome[meio] != ' ': meio -= 1
    if meio == 0: return nome[:max_chars], nome[max_chars:]
    return nome[:meio], nome[meio+1:]

def gerar_etiqueta(c):
    W,H = 439,280
    nome=zpl_safe(c['nome']); sexo=c.get('sexo','M')
    tribo=zpl_safe(c.get('tribo','')); cel=zpl_safe(c.get('celular',''))
    cid=zpl_safe(c.get('cidade','')); cor=TRIBO_COR_NOME.get(c.get('tribo',''),'')
    fc=c.get('fc',0)
    n1,n2=quebra_nome(nome,18)
    z = f"^XA\n^PW{W}\n^LL{H}\n^LH0,0\n"
    z += f"^FO3,3^GB433,274,3^FS\n"
    z += f"^FO310,6^A0N,18,18^FDFicha: {fc:03d}^FS\n"
    if n2:
        z += f"^FO8,8^ADN,36,20^FD{n1}^FS\n"
        z += f"^FO8,46^ADN,34,18^FD{n2} ({sexo})^FS\n"
        y1=84
    else:
        z += f"^FO8,8^ADN,36,20^FD{n1} ({sexo})^FS\n"
        y1=50
    z += f"^FO6,{y1}^GB427,2,2^FS\n"
    y_tel=y1+5; y_cid=y_tel+22
    z += f"^FO8,{y_tel}^A0N,19,19^FDTel: {cel}^FS\n"
    z += f"^FO8,{y_cid}^A0N,19,19^FDCidade: {cid}^FS\n"
    y2=y_cid+24
    z += f"^FO6,{y2}^GB427,2,2^FS\n"
    y_tri=y2+3
    z += f"^FO6,{y_tri}^GB200,40,40^FS\n"
    z += f"^FO10,{y_tri+4}^ADN,32,18^FR^FD{tribo.upper()}^FS\n"
    z += f"^FO212,{y_tri+8}^ADN,24,14^FDCor: {cor}^FS\n"
    y_vol=y_tri+46
    if y_vol+20 < H-4:
        z += f"^FO8,{y_vol}^A0N,19,19^FDVolumes: _____________________^FS\n"
    z += "^XZ\n"
    return z

def gerar_zpl_campistas(campistas, copias=4):
    out = ''
    for c in sorted(campistas, key=lambda x: x.get('nome','').lower()):
        for _ in range(copias):
            out += gerar_etiqueta(c)
    return out

# ── Multipart parser ──
def parse_multipart(content_type, body):
    import re
    boundary = None
    for part in content_type.split(';'):
        part = part.strip()
        if part.startswith('boundary='):
            boundary = part[9:].strip('"'); break
    if not boundary: return {}, {}
    files, fields = {}, {}
    delimiter = ('--' + boundary).encode()
    parts = body.split(delimiter)
    for part in parts[1:]:
        if part.strip() in (b'', b'--', b'--\r\n'): continue
        if part.startswith(b'--'): continue
        if b'\r\n\r\n' in part: hraw, content = part.split(b'\r\n\r\n', 1)
        elif b'\n\n' in part:   hraw, content = part.split(b'\n\n', 1)
        else: continue
        content = content.rstrip(b'\r\n')
        hs = hraw.decode('utf-8', errors='replace')
        cd = re.search(r'Content-Disposition:[^\r\n]*name="([^"]+)"', hs, re.I)
        if not cd: continue
        name = cd.group(1)
        fn = re.search(r'filename="([^"]*)"', hs, re.I)
        if fn: files[name] = content
        else:  fields[name] = content.decode('utf-8', errors='replace')
    return files, fields


class Handler(BaseHTTPRequestHandler):

    def log_message(self, fmt, *args):
        print(f"  {self.address_string()} {fmt % args}")

    def _cookies(self):
        c = {}
        for part in self.headers.get('Cookie','').split(';'):
            if '=' in part:
                k,v = part.strip().split('=',1); c[k.strip()]=v.strip()
        return c

    def _authed(self):
        return sessao_valida(self._cookies().get('session'))

    def _json(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False).encode('utf-8')
        self.send_response(status)
        self.send_header('Content-Type','application/json; charset=utf-8')
        self.send_header('Content-Length', len(body))
        self.end_headers(); self.wfile.write(body)

    def _cors(self):
        self.send_header('Access-Control-Allow-Origin','*')

    def _404(self):
        self.send_response(404); self.end_headers(); self.wfile.write(b'Not found')

    def do_OPTIONS(self):
        self.send_response(200); self._cors(); self.end_headers()

    def do_GET(self):
        path = urlparse(self.path).path
        if path in ('/', '/index.html'):
            if self._authed():
                self._serve_file('gerador_tribos.html', 'text/html; charset=utf-8')
            else:
                self._serve_login()
        elif path == '/api/restricoes':
            if not self._authed(): self._json({'error':'Não autorizado'},401); return
            self._json(load_restricoes())
        else:
            self._404()

    def do_POST(self):
        if self.path == '/login':
            self._handle_login()
        elif self.path == '/logout':
            t = self._cookies().get('session')
            if t in SESSOES: del SESSOES[t]
            self.send_response(302)
            self.send_header('Location','/')
            self.send_header('Set-Cookie','session=; Max-Age=0; Path=/')
            self.end_headers()
        elif self.path == '/api/atualizar':
            if not self._authed(): self._json({'error':'Não autorizado'},401); return
            self._handle_atualizar()
        elif self.path == '/api/conflito':
            if not self._authed(): self._json({'error':'Não autorizado'},401); return
            self._handle_conflito()
        elif self.path == '/api/del_conflito':
            if not self._authed(): self._json({'error':'Não autorizado'},401); return
            self._handle_del_conflito()
        elif self.path == '/api/etiquetas':
            if not self._authed(): self._json({'error':'Não autorizado'},401); return
            self._handle_etiquetas()
        else:
            self._404()

    def _handle_login(self):
        length = int(self.headers.get('Content-Length',0))
        body = self.rfile.read(length)
        try: data = json.loads(body); senha = data.get('senha','')
        except: senha = ''
        if senha == SENHA:
            token = nova_sessao()
            self.send_response(200)
            self.send_header('Content-Type','application/json')
            self.send_header('Set-Cookie',f'session={token}; Max-Age=28800; Path=/; HttpOnly')
            self.end_headers()
            self.wfile.write(json.dumps({'ok':True}).encode())
        else:
            self._json({'ok':False,'error':'Senha incorreta'},401)

    def _handle_atualizar(self):
        if not os.path.exists(MODELO):
            self._json({'error':f'Modelo VIII_MODELO.xlsx não encontrado em {DIR}'},500); return
        ct = self.headers.get('Content-Type','')
        length = int(self.headers.get('Content-Length',0))
        raw = self.rfile.read(length)
        files, fields = parse_multipart(ct, raw)
        csv_data = files.get('csv')
        if not csv_data:
            self._json({'error':'Arquivo CSV não enviado.'},400); return
        ext = fields.get('filename','inscricoes.csv').split('.')[-1].lower()
        ext = ext if ext in ['csv','xlsx','xls','xlsm'] else 'csv'

        with tempfile.NamedTemporaryFile(suffix='.'+ext, delete=False) as f:
            f.write(csv_data); csv_path = f.name
        out_path = tempfile.mktemp(suffix='.xlsx')
        try:
            from atualizar_tribos import atualizar
            result = atualizar(csv_path, out_path)
            if not os.path.exists(out_path):
                self._json({'error':'Planilha não foi gerada.'},500); return
            # Copia como novo modelo
            shutil.copy2(out_path, MODELO)
            with open(out_path,'rb') as f: xlsx_b64 = base64.b64encode(f.read()).decode()
            # Lê resultado para retornar stats
            import openpyxl
            wb = openpyxl.load_workbook(out_path)
            ws = wb['Cadastro Geral Campistas']
            hdr = {cell.value: j for j,cell in enumerate(ws[1],1) if cell.value}
            campistas = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                try: fc=int(row[0])
                except: continue
                nome = str(row[hdr.get('NOME',2)-1] or '').strip()
                if nome:
                    campistas.append({
                        'fc': fc, 'nome': nome,
                        'tribo': str(row[hdr.get('TRIBO',17)-1] or ''),
                        'sexo': str(row[hdr.get('SEXO',4)-1] or '')[0].upper() if row[hdr.get('SEXO',4)-1] else 'M',
                        'celular': str(row[hdr.get('CELULAR',8)-1] or ''),
                        'cidade': str(row[hdr.get('CIDADE',9)-1] or ''),
                        '_novo': str(row[hdr.get('TRIBO',17)-1] or '') != ''
                    })
            self._json({
                'ok': True,
                'n_campistas': len(campistas),
                'xlsx_b64': xlsx_b64,
                'campistas': campistas,
                'avisos': result.get('avisos',[]) if isinstance(result,dict) else [],
                'erros':  result.get('erros',[])  if isinstance(result,dict) else [],
            })
        except Exception as e:
            traceback.print_exc()
            self._json({'error': str(e)},500)
        finally:
            for p in [csv_path, out_path]:
                if p and os.path.exists(p):
                    try: os.unlink(p)
                    except: pass

    def _handle_conflito(self):
        length = int(self.headers.get('Content-Length',0))
        body = self.rfile.read(length)
        try: data = json.loads(body)
        except: self._json({'error':'JSON inválido'},400); return
        rest = load_restricoes()
        rest['restricoes'].append(data)
        save_restricoes(rest)
        self._json({'ok':True, 'total': len(rest['restricoes'])})

    def _handle_del_conflito(self):
        length = int(self.headers.get('Content-Length',0))
        body = self.rfile.read(length)
        try: data = json.loads(body); idx = int(data.get('idx',-1))
        except: self._json({'error':'JSON inválido'},400); return
        rest = load_restricoes()
        if 0 <= idx < len(rest['restricoes']):
            removed = rest['restricoes'].pop(idx)
            save_restricoes(rest)
            self._json({'ok':True,'removed':removed})
        else:
            self._json({'error':'Índice inválido'},400)

    def _handle_etiquetas(self):
        length = int(self.headers.get('Content-Length',0))
        body = self.rfile.read(length)
        try: data = json.loads(body)
        except: self._json({'error':'JSON inválido'},400); return
        campistas = data.get('campistas',[])
        copias    = int(data.get('copias',4))
        if not campistas:
            self._json({'error':'Nenhum campista enviado'},400); return
        zpl = gerar_zpl_campistas(campistas, copias)
        zpl_b64 = base64.b64encode(zpl.encode('utf-8')).decode()
        self._json({'ok':True,'zpl_b64':zpl_b64,'n_etiquetas':len(campistas)*copias})

    def _serve_login(self):
        html = open(os.path.join(DIR,'login.html'),'rb').read() if os.path.exists(os.path.join(DIR,'login.html')) else self._login_inline()
        self.send_response(200)
        self.send_header('Content-Type','text/html; charset=utf-8')
        self.send_header('Content-Length',len(html))
        self.end_headers(); self.wfile.write(html)

    def _login_inline(self):
        return '''<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8"><title>Login</title>
<style>*{margin:0;padding:0;box-sizing:border-box}body{font-family:Georgia,serif;background:#FAF6EE;display:flex;align-items:center;justify-content:center;min-height:100vh}.card{background:white;border-radius:16px;padding:40px;width:360px;box-shadow:0 4px 24px rgba(0,0,0,.08);text-align:center}h1{font-size:22px;margin-bottom:6px}.sub{font-size:13px;color:#888;font-style:italic;margin-bottom:28px}input{width:100%;padding:12px;border:1.5px solid #ddd;border-radius:8px;font-size:15px;margin-bottom:16px;outline:none}input:focus{border-color:#C8962A}button{width:100%;padding:13px;background:linear-gradient(135deg,#C8962A,#FF4500);color:white;border:none;border-radius:50px;font-size:14px;font-weight:600;cursor:pointer}.err{color:#C0392B;font-size:13px;margin-top:10px;display:none}</style></head>
<body><div class="card"><h1>Guardioes</h1><p class="sub">Acampamento Espírito Empreendedor</p>
<input type="password" id="s" placeholder="Senha de acesso" onkeydown="if(event.key==='Enter')login()">
<button onclick="login()">Entrar</button><div class="err" id="e">Senha incorreta.</div></div>
<script>async function login(){const r=await fetch('/login',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({senha:document.getElementById('s').value})});const d=await r.json();if(d.ok)location.href='/';else document.getElementById('e').style.display='block';}</script></body></html>'''.encode('utf-8')

    def _serve_file(self, filename, ctype):
        path = os.path.join(DIR, filename)
        try:
            with open(path,'rb') as f: data = f.read()
            self.send_response(200)
            self.send_header('Content-Type',ctype)
            self.send_header('Content-Length',len(data))
            self.end_headers(); self.wfile.write(data)
        except FileNotFoundError:
            self._404()


if __name__ == '__main__':
    os.chdir(DIR)
    print("=" * 52)
    print("  Sistema Guardiões do Amor Maior")
    print("=" * 52)
    print(f"\n🔒  Senha: {SENHA}")
    print(f"🌐  http://localhost:{PORT}")
    print(f"📋  Modelo: {MODELO}")
    print(f"   ({'✅ existe' if os.path.exists(MODELO) else '❌ NÃO ENCONTRADO'})")
    print("   (Ctrl+C para parar)\n")
    server = HTTPServer(('', PORT), Handler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print('\nServidor encerrado.')
