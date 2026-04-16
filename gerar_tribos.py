"""
gerar_tribos.py
===============
Uso: python3 gerar_tribos.py <arquivo_base.xlsx>

Recebe uma planilha com a aba "Cadastro Geral Campistas" contendo as colunas:
  A=FICHA  B=NOME  C=CAMISETA  D=SEXO  E=PESO  F=ALTURA  G=IDADE
  H=CELULAR  I=CIDADE  J=CONFLITOS  O=ALERGIAS  R=PARCEIRO FC

E gera a planilha completa com:
  1. Cadastro Geral Campistas (base enriquecida com TRIBO e PARCEIRO FC)
  2. Relatorio Familia Provas (formato idêntico ao editado, VLOOKUPs, alertas)
  3. Relatorio Chamada Onibus
  4. CONFLITO CONHECIDOS

Critérios automáticos:
  - 9 tribos × 8 homens + 8 mulheres
  - Casais/parceiros em tribos diferentes
  - Balanceamento de peso e idade (snake draft)
  - VLOOKUP em todas as células de dados (troca pelo FC)
  - Alertas: cônjuge na mesma tribo 🟠, FC inválido 🔴, 8F/8M verde/vermelho
"""

import sys, re
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────
TRIBOS       = ['Simeão','Rubem','Judá','Levi','Benjamim','Issacar','Gad','Zabulom','EFRAIM']
CORES        = ['FFC000','FF0000','FF6600','FF99CC','0070C0','00B050','8B4513','808080','BFBFBF']
CORES_LIGHT  = ['FFF2CC','FFE8E8','FFE5CC','FFD6EC','DAEEF3','EAF6E2','F5E6DA','E8E8E8','F5F5F5']
CORES_NOME   = ['Amarela','Vermelha','Laranja','Rosa','Azul','Verde','Marrom','Cinza','Cinza Claro']
TEXTO_COR    = ['000000','FFFFFF','FFFFFF','000000','FFFFFF','FFFFFF','FFFFFF','FFFFFF','000000']
DINAMICAS    = ['Casa do Pai','Espelho','Balanço','Casa de Maria','Salto da Fé',
                'Cabana','Casa da Misericórdia','Metrô','Paredão']

REF   = "'Cadastro Geral Campistas'!$A:$R"
# Índices das colunas no Cadastro Geral (1-based para VLOOKUP)
CI    = dict(ficha=1,nome=2,camiseta=3,sexo=4,peso=5,altura=6,idade=7,
             celular=8,cidade=9,conflitos=10,c1n=11,c1t=12,c2n=13,c2t=14,
             alergias=15,nasc=16,tribo=17,parceiro=18)

ANJOS_PADRAO = {t: {'a1':'(preencher)','a2':'(preencher)','g':'(preencher)'} for t in TRIBOS}

# ─────────────────────────────────────────────
# HELPERS DE ESTILO
# ─────────────────────────────────────────────
def F(bold=False,size=10,color='000000',italic=False):
    return Font(bold=bold,size=size,name='Arial',color=color,italic=italic)
def P(color): return PatternFill('solid',fgColor=color)
def A(h='left',v='center',wrap=False): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def B(style='thin'):
    s=Side(style=style)
    return Border(left=s,right=s,top=s,bottom=s)
def BM(): # medium left, thin others
    return Border(left=Side(style='medium'),right=Side(style='thin'),
                  top=Side(style='thin'),bottom=Side(style='thin'))
def BNONE(): return Border()

def set_cell(ws,row,col,val,font=None,fill=None,align=None,border=None):
    c=ws.cell(row=row,column=col,value=val)
    if font:   c.font=font
    if fill:   c.fill=fill
    if align:  c.alignment=align
    if border: c.border=border
    return c

def is_skip(s):
    return str(s).lower().strip() in [
        'não','nao','não.','nao.','não!','não que eu saiba',
        'nao que eu saiba','','nan','0','não que saiba','none','false']

# ─────────────────────────────────────────────
# 1. LEITURA DOS DADOS
# Aceita duas origens:
#   a) ler_inscricoes(path) — planilha original de inscrições, gera FC 1..N
#   b) ler_base(path)       — planilha já processada com aba Cadastro Geral
# ─────────────────────────────────────────────
from datetime import datetime as _dt

def _clean_altura(v):
    try:
        s = str(v).replace(',','.').replace('m','').strip()
        f = float(s)
        return round(f/100 if f > 3 else f, 2)
    except Exception:
        return None

def _clean_peso(v):
    try:
        s = str(v).replace('kg','').replace('Kg','').strip()
        return float(s)
    except Exception:
        return None

def _calc_idade(v, ref=None):
    if ref is None: ref = _dt.today()
    try:
        if v is None or str(v).strip() in ('','nan','NaT','None'): return None
        return int((_dt.today() - pd.to_datetime(v)).days / 365.25)
    except Exception:
        return None

def _clean_sexo(v):
    return 'M' if str(v).strip().lower().startswith('m') else 'F'

def _s(v, default=''):
    if v is None or str(v).strip() in ('','nan','None','NaT','NaN'): return default
    return str(v).strip()

def _col(df_cols, candidates):
    for c in candidates:
        if c in df_cols: return c
    # Busca case-insensitive como fallback
    lower_map = {col.lower(): col for col in df_cols}
    for c in candidates:
        if c.lower() in lower_map: return lower_map[c.lower()]
    return None


def ler_inscricoes(path):
    """
    Le planilha de inscricoes original (exportacao do sistema).
    Aceita .xlsx, .xls, .xlsm e .csv (qualquer encoding/separador).
    Gera FC de 1..N em ordem alfabetica por nome.
    Ignora inscricoes com Cancelada? = Sim.
    """
    import os, csv as _csv
    ext = os.path.splitext(str(path))[-1].lower().lstrip('.')

    if ext == 'csv':
        # Detecta encoding
        try:
            import chardet
            with open(path, 'rb') as f:
                raw = f.read(8192)
            enc = chardet.detect(raw).get('encoding') or 'utf-8'
        except ImportError:
            enc = 'utf-8'

        # Tenta leitura com varios encodings e separadores
        df_raw = None
        for encoding in [enc, 'utf-8-sig', 'latin-1', 'cp1252', 'utf-8']:
            for sep in [',', ';', '\t', '|']:
                try:
                    df_raw = pd.read_csv(path, encoding=encoding, sep=sep,
                                         on_bad_lines='skip', dtype=str)
                    if len(df_raw.columns) > 3:
                        print(f"   CSV lido: encoding={encoding} separador='{sep}'")
                        break
                    df_raw = None
                except Exception:
                    continue
            if df_raw is not None:
                break
        if df_raw is None:
            raise ValueError("Nao foi possivel ler o CSV. Tente salvar como UTF-8 no Excel.")
    else:
        df_raw = pd.read_excel(path, header=0, dtype=str)

    df_raw.columns = [c.strip() for c in df_raw.columns]
    cols = list(df_raw.columns)

    # Filtra cancelados (Cancelada? = Sim)
    c_cancel = _col(cols, ['Cancelada?', 'Cancelada', 'Cancelado'])
    if c_cancel:
        mask = df_raw[c_cancel].astype(str).str.strip().str.lower() == 'sim'
        n_cancel = int(mask.sum())
        if n_cancel > 0:
            print(f"   {n_cancel} inscricao(oes) cancelada(s) ignorada(s)")
        df_raw = df_raw[~mask].reset_index(drop=True)

    # Mapeamento de colunas
    C = {
        'nome':   _col(cols, ['Nome']),
        'sexo':   _col(cols, ['Sexo']),
        'camis':  _col(cols, ['Tamanho da Camiseta', 'Tamanho Camiseta', 'Camiseta']),
        'peso':   _col(cols, ['Peso']),
        'altura': _col(cols, ['Altura ', 'Altura']),
        'nasc':   _col(cols, ['Data de Nascimento', 'Nascimento']),
        'cel':    _col(cols, ['Celular', 'Telefone']),
        'cidade': _col(cols, ['Cidade']),
        'conf':   _col(cols, [
            'Tem algum Familiar ou amigo que ira fazer o acampamento? Se sim nos indique o nome por favor.',
            'Tem algum Familiar ou amigo que ir\u00e1 fazer o acampamento? Se sim nos indique o nome por favor.',
            'Familiar ou amigo', 'Conflitos', 'Conhecido']),
        'c1n':    _col(cols, [
            'Contato 1: Nos indique o nome de um amigo ou familiar para contato em caso de necessidade:',
            'Contato 1: N\u00f3s indique o nome de um amigo ou familiar para contato em caso de necessidade:',
            'Contato 1 Nome', 'Nomes Contatos I']),
        'c1t':    _col(cols, ['Telefone Contato 1', 'Telefone Contato I']),
        'c2n':    _col(cols, [
            'Contato 2: Nos indique o nome de um amigo ou familiar para contato em caso de necessidade:',
            'Contato 2: N\u00f3s indique o nome de um amigo ou familiar para contato em caso de necessidade:',
            'Contato 2 Nome', 'Nomes Contatos II']),
        'c2t':    _col(cols, ['Telefone Contato 2', 'Telefone Contato II']),
        'alerg':  _col(cols, [
            'Possui alguma alergia a alimentos, medicamentos, insetos etc? (liste todas que possuir):',
            'Alergias', 'Alergia']),
        'cupom':  _col(cols, ['Cupom']),
        'categ':  _col(cols, ['Categoria']),
        'lider':  _col(cols, ['Nome da inscricao lider',
                               'Nome da inscri\u00e7\u00e3o l\u00edder', 'Lider']),
        'doc':    _col(cols, ['Documento']),
    }

    def gv(row, key, default=''):
        colname = C.get(key)
        if colname is None:
            return default
        v = row.get(colname)
        return default if v is None or str(v).strip() in ('', 'nan', 'None', 'NaT', 'NaN') else str(v).strip()

    rows = []
    for _, r in df_raw.iterrows():
        nome = _s(r.get(C['nome'], '') if C['nome'] else '')
        if not nome:
            continue

        nasc_raw = r.get(C['nasc']) if C['nasc'] else None
        nasc_str = ''
        try:
            nasc_str = pd.to_datetime(nasc_raw).strftime('%Y-%m-%d')
        except Exception:
            nasc_str = _s(nasc_raw)

        rows.append({
            'Nome':      nome,
            'Sexo':      _clean_sexo(gv(r, 'sexo', 'M')),
            'Camiseta':  gv(r, 'camis'),
            'Peso':      _clean_peso(r.get(C['peso']) if C['peso'] else None),
            'Altura':    _clean_altura(r.get(C['altura']) if C['altura'] else None),
            'Idade':     _calc_idade(r.get(C['nasc']) if C['nasc'] else None),
            'Celular':   gv(r, 'cel'),
            'Cidade':    gv(r, 'cidade'),
            'Conflitos': gv(r, 'conf'),
            'C1Nome':    gv(r, 'c1n'),
            'C1Tel':     gv(r, 'c1t'),
            'C2Nome':    gv(r, 'c2n'),
            'C2Tel':     gv(r, 'c2t'),
            'Alergias':  gv(r, 'alerg'),
            'Nasc':      nasc_str,
            'ParcFC':    '',
            'DocNum':    gv(r, 'doc'),
            '_cupom':    gv(r, 'cupom').upper(),
            '_categ':    gv(r, 'categ').upper(),
            '_lider':    gv(r, 'lider'),
        })

    # Ordena por nome e gera FC 1..N
    rows.sort(key=lambda x: x['Nome'].lower())
    for i, r in enumerate(rows, 1):
        r['FC'] = i

    df = pd.DataFrame(rows)
    df['Peso_N']  = pd.to_numeric(df['Peso'],  errors='coerce')
    df['Idade_N'] = pd.to_numeric(df['Idade'], errors='coerce')
    df['Alt_N']   = pd.to_numeric(df['Altura'], errors='coerce')

    m = (df['Sexo'] == 'M').sum()
    f = (df['Sexo'] == 'F').sum()
    print(f"   {len(df)} campistas (M:{m} F:{f}) — FC 1 a {len(df)}")
    return df

def ler_base(path):
    """Lê planilha base ja processada com aba Cadastro Geral Campistas."""
    wb = load_workbook(path)
    ws = wb['Cadastro Geral Campistas']
    header = {cell.value: j for j, cell in enumerate(ws[1], 1) if cell.value}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        try: int(row[0])
        except (ValueError, TypeError): continue

        def g(nome, default=''):
            idx = header.get(nome)
            return default if idx is None or row[idx-1] is None else row[idx-1]

        rows.append({
            'FC':       int(g('FICHA',0)),
            'Nome':     _s(g('NOME')),
            'Camiseta': _s(g('TAMANHO DA CAMISETA')),
            'Sexo':     _clean_sexo(g('SEXO','M')),
            'Peso':     _clean_peso(g('PESO')),
            'Altura':   _clean_altura(g('ALTURA')),
            'Idade':    _calc_idade(g('DATA DE NASCIMENTO')),
            'Celular':  _s(g('CELULAR')),
            'Cidade':   _s(g('CIDADE')),
            'Conflitos':_s(g('CONFLITOS / CONHECIDOS')),
            'C1Nome':   _s(g('NOMES CONTATOS I')),
            'C1Tel':    _s(g('TELEFONE CONTATO I')),
            'C2Nome':   _s(g('NOMES CONTATOS II')),
            'C2Tel':    _s(g('TELEFONE CONTATO II')),
            'Alergias': _s(g('ALERGIAS')),
            'Nasc':     _s(g('DATA DE NASCIMENTO')).split(' ')[0],
            'TRIBO':    _s(g('TRIBO', '')),
            'ParcFC':   _s(g('PARCEIRO FC','')),
            '_cupom':   '',
            '_categ':   '',
            '_lider':   '',
        })

    df = pd.DataFrame(rows)
    df['Peso_N']  = pd.to_numeric(df['Peso'],  errors='coerce')
    df['Idade_N'] = pd.to_numeric(df['Idade'], errors='coerce')
    df['Alt_N']   = pd.to_numeric(df['Altura'],errors='coerce')
    return df


# 2. IDENTIFICAR CASAIS
# Fontes (em ordem de prioridade):
#   1. Coluna PARCEIRO FC (ja preenchida na base)
#   2. Cupom contendo CASAL + coluna lider da inscricao
#   3. Categoria contendo CASAL/DIFERENCA/LINKPAG + lider
#   4. Conflitos com palavras esposo/esposa/marido
# ─────────────────────────────────────────────
def identificar_casais(df, df_inscricoes=None):
    partner = {}

    def add_par(fc1, fc2):
        if fc1 and fc2 and fc1 != fc2:
            partner[fc1] = fc2
            partner[fc2] = fc1

    # Fonte 1: PARCEIRO FC ja na base
    for _, r in df.iterrows():
        p = r['ParcFC']
        if p and str(p).strip() not in ['', 'nan', '0']:
            try:
                add_par(int(r['FC']), int(p))
            except Exception:
                pass

    # Fontes 2 e 3: Cupom/Categoria — usa campos _cupom/_categ/_lider do df
    # (preenchidos por ler_inscricoes) ou da planilha de inscricoes separada
    nome_to_fc_idx = {r['Nome'].strip().lower(): int(r['FC']) for _, r in df.iterrows()}

    def nome_para_fc(nome_str):
        if not nome_str or str(nome_str).strip() in ['', 'nan']:
            return None
        alvo = str(nome_str).strip().lower()
        if alvo in nome_to_fc_idx:
            return nome_to_fc_idx[alvo]
        for nb, fc in nome_to_fc_idx.items():
            if alvo in nb or nb in alvo:
                return fc
        return None

    # Usa campos embutidos no df (vindos de ler_inscricoes)
    if '_cupom' in df.columns:
        for _, r in df.iterrows():
            cupom    = str(r.get('_cupom','')).upper()
            categ    = str(r.get('_categ','')).upper()
            lider_nm = str(r.get('_lider','')).strip()
            is_casal = ('CASAL' in cupom) or any(
                c in categ for c in ['CASAL','DIFERENCA','LINKPAG'])
            if is_casal and lider_nm and lider_nm not in ['','nan']:
                fc1 = int(r['FC'])
                fc2 = nome_para_fc(lider_nm)
                if fc1 and fc2:
                    add_par(fc1, fc2)
                    print(f"   [Cupom/Cat] FC{fc1} {r['Nome']} <-> FC{fc2} {lider_nm}")

    # Planilha de inscricoes separada (compatibilidade com modo anterior)
    if df_inscricoes is not None:
        for _, r in df_inscricoes.iterrows():
            nome_r   = str(r.get('Nome', '')).strip()
            cupom    = str(r.get('Cupom', '')).upper()
            categ    = str(r.get('Categoria', '')).upper()
            lider_nm = str(r.get('Nome da inscrição líder',
                           r.get('Nome da inscricao lider', ''))).strip()
            is_casal = ('CASAL' in cupom) or any(
                c in categ for c in ['CASAL','DIFERENCA','LINKPAG'])
            if is_casal and lider_nm and lider_nm not in ['','nan']:
                fc1 = nome_para_fc(nome_r)
                fc2 = nome_para_fc(lider_nm)
                if fc1 and fc2:
                    add_par(fc1, fc2)
                    print(f"   [Cupom/Cat] FC{fc1} {nome_r} <-> FC{fc2} {lider_nm}")

    # Fonte 4: Conflitos com esposo/esposa/marido
    nome_to_fc2 = {r['Nome'].strip().lower(): int(r['FC']) for _, r in df.iterrows()}
    palavras = ['esposo', 'esposa', 'marido', 'minha esposa', 'meu esposo', 'meu marido']
    for _, r in df.iterrows():
        conf = str(r['Conflitos']).lower()
        if not any(w in conf for w in palavras):
            continue
        fc1 = int(r['FC'])
        if fc1 in partner:
            continue
        for nome_base, fc2 in nome_to_fc2.items():
            if fc2 != fc1 and nome_base in conf:
                add_par(fc1, fc2)
                break

    return partner

# ─────────────────────────────────────────────
# 3. BALANCEAMENTO SNAKE DRAFT
# ─────────────────────────────────────────────
def normalizar(s):
    mn,mx = s.min(),s.max()
    return (s-mn)/(mx-mn) if mx>mn else pd.Series(0.0,index=s.index)

def atribuir_tribos(df, partner):
    n = len(TRIBOS)
    df = df.copy()
    df['Score'] = normalizar(df['Peso_N'].fillna(df['Peso_N'].median())) * 0.5 + \
                  normalizar(df['Idade_N'].fillna(df['Idade_N'].median())) * 0.5

    def snake(grupo_df, other_assign=None):
        gs = grupo_df.sort_values('Score').reset_index(drop=True)
        assign = {}
        counts = [0]*n
        i = 0; rnd = 0
        while i < len(gs):
            order = list(range(n)) if rnd%2==0 else list(range(n-1,-1,-1))
            for ti in order:
                if i >= len(gs): break
                fc  = int(gs.iloc[i]['FC'])
                p   = partner.get(fc)
                t   = TRIBOS[ti]
                p_t = (other_assign or assign).get(p)
                if p_t == t:
                    # busca próxima tribo sem conflito
                    placed = False
                    for alt in order[order.index(ti)+1:]+order[:order.index(ti)]:
                        at = TRIBOS[alt]
                        if (other_assign or assign).get(p) != at:
                            assign[fc]=at; counts[alt]+=1; placed=True; break
                    if not placed:
                        assign[fc]=t; counts[ti]+=1
                else:
                    assign[fc]=t; counts[ti]+=1
                i+=1
            rnd+=1
        return assign

    f_df = df[df['Sexo']=='F']
    m_df = df[df['Sexo']=='M']
    af   = snake(f_df)
    am   = snake(m_df, other_assign=af)
    full = {**af, **am}

    # Corrige contagens 8/8
    for t in TRIBOS:
        for sexo, target in [('M',8),('F',8)]:
            grupo = df[(df['TRIBO_NEW']==t) & (df['Sexo']==sexo)] if 'TRIBO_NEW' in df else pd.DataFrame()
            pass

    df['TRIBO'] = df['FC'].apply(lambda fc: full.get(int(fc),'?'))

    # Corrige desequilíbrios — tolera M≠F quando inscrições são desbalanceadas
    import math as _math
    for sexo in ['M', 'F']:
        total_s = len(df[df['Sexo']==sexo])
        per_t   = _math.floor(total_s / len(TRIBOS))
        extra   = total_s % len(TRIBOS)
        max_for = {TRIBOS[i]: per_t+1 if i<extra else per_t for i in range(len(TRIBOS))}
        iters = 0
        while iters < 50:
            moved = False
            for t in TRIBOS:
                g = df[(df['TRIBO']==t)&(df['Sexo']==sexo)]
                if len(g) <= max_for[t]: continue
                excess_fc = int(g.iloc[-1]['FC'])
                p_fc = partner.get(excess_fc)
                for t2 in TRIBOS:
                    if t2 == t: continue
                    g2 = df[(df['TRIBO']==t2)&(df['Sexo']==sexo)]
                    p_t2 = None
                    if p_fc and len(df[df['FC']==p_fc]) > 0:
                        p_t2 = df[df['FC']==p_fc]['TRIBO'].values[0]
                    if len(g2) < max_for[t2] and p_t2 != t2:
                        df.loc[df['FC']==excess_fc, 'TRIBO'] = t2
                        moved = True; break
                if moved: break
            if not moved: break
            iters += 1
    return df

# ─────────────────────────────────────────────
# 4. GERAR PLANILHA
# ─────────────────────────────────────────────
def gerar(df, partner, anjos=None, output_path='output.xlsx'):
    if anjos is None: anjos = ANJOS_PADRAO
    wb = Workbook()

    global_peso  = round(df['Peso_N'].mean(),1)
    global_idade = round(df['Idade_N'].mean(),1)

    # ── VLOOKUP helper ──
    def vl(ref, col): return f"=IFERROR(VLOOKUP({ref},{REF},{col},0),\"\")"

    # ── Fórmula alerta col I ──
    def alert_formula(afc, fc_range):
        p    = f"IFERROR(VLOOKUP({afc},{REF},{CI['parceiro']},0),0)"
        conf = f"AND({p}>0,COUNTIF({fc_range},{p})>0)"
        pn   = f"IFERROR(VLOOKUP({p},{REF},{CI['nome']},\"\"),\"\")"
        cv   = f"IFERROR(VLOOKUP({afc},{REF},{CI['conflitos']},0),\"\")"
        av   = f"IFERROR(VLOOKUP({afc},{REF},{CI['alergias']},0),\"\")"
        ok_c = (f"AND({cv}<>\"\",LOWER(TRIM({cv}))<>\"não\","
                f"LOWER(TRIM({cv}))<>\"nao\",LOWER(TRIM({cv}))<>\"0\")")
        ok_a = (f"AND({av}<>\"\",LOWER(TRIM({av}))<>\"não\","
                f"LOWER(TRIM({av}))<>\"nao\",LOWER(TRIM({av}))<>\"0\")")
        obs  = (f"TRIM(IF({ok_c},{cv}&IF({ok_a},\" | Alergia: \"&{av},\"\"),"
                f"IF({ok_a},\"Alergia: \"&{av},\"\")))")
        return (f"=IF({conf},\"⚠️ CONFLITO! Cônjuge na mesma tribo: \"&{pn},{obs})")

    # ── CF helpers ──
    def cf_expr(ws, rng, formula, fill_hex, font_hex='FFFFFF', priority=1):
        ds = DifferentialStyle(
            fill=PatternFill(start_color=fill_hex,end_color=fill_hex,fill_type='solid'),
            font=Font(bold=True,color=font_hex,name='Arial',size=10))
        ws.conditional_formatting.add(rng, Rule(type='expression',dxf=ds,priority=priority,formula=[formula]))

    # ════════════════════════════
    # ABA 1 — Cadastro Geral Campistas
    # ════════════════════════════
    ws1 = wb.active; ws1.title = 'Cadastro Geral Campistas'
    hdrs=['FICHA','NOME','TAMANHO DA CAMISETA','SEXO','PESO','ALTURA','IDADE',
          'CELULAR','CIDADE','CONFLITOS / CONHECIDOS','NOMES CONTATOS I','TELEFONE CONTATO I',
          'NOMES CONTATOS II','TELEFONE CONTATO II','ALERGIAS','DATA DE NASCIMENTO','TRIBO','PARCEIRO FC']
    cw1 =[7,40,10,6,8,8,7,22,20,45,25,18,25,18,30,18,12,11]
    for j,h in enumerate(hdrs,1):
        set_cell(ws1,1,j,h,font=F(bold=True,color='FFFFFF'),fill=P('1F4E79'),
                 align=A('center'),border=B())
    ws1.row_dimensions[1].height=20

    df_sorted = df.sort_values('FC').reset_index(drop=True)
    for i,(_,r) in enumerate(df_sorted.iterrows(),2):
        ti  = TRIBOS.index(r['TRIBO'])
        ic  = int(r['FC']) in partner
        parc= partner.get(int(r['FC']),'')
        rf  = P('FFF0CC') if ic else (P('DCE6F1') if i%2==0 else P('FFFFFF'))
        vals=[r['FC'],r['Nome'],r['Camiseta'],r['Sexo'],r['Peso'],r['Altura'],
              r['Idade'],r['Celular'],r['Cidade'],r['Conflitos'],r['C1Nome'],r['C1Tel'],
              r['C2Nome'],r['C2Tel'],r['Alergias'],r['Nasc'],r['TRIBO'],parc]
        for j,v in enumerate(vals,1):
            c=set_cell(ws1,i,j,v,font=F(),fill=rf,
                       align=A('center' if j in [1,3,4,5,6,7,18] else 'left'),border=B())
        ws1.cell(row=i,column=17).fill=P(CORES[ti])
        ws1.cell(row=i,column=17).font=F(bold=True,color=TEXTO_COR[ti])
        if parc: ws1.cell(row=i,column=18).font=F(bold=True,color='8B4513')
    for j,w in enumerate(cw1,1): ws1.column_dimensions[get_column_letter(j)].width=w
    ws1.freeze_panes='A2'
    nota=ws1.cell(row=len(df_sorted)+3,column=1,
                  value='⚠️ Esta aba é a BASE de todos os dados. Não exclua linhas.')
    nota.font=Font(bold=True,color='C00000',size=10,name='Arial')
    ws1.merge_cells(f'A{len(df_sorted)+3}:R{len(df_sorted)+3}')

    # ════════════════════════════
    # ABA 2 — Relatorio Familia Provas
    # ════════════════════════════
    ws2 = wb.create_sheet('Relatorio Familia Provas')
    # Larguras EXATAS da planilha editada
    for col,w in zip('ABCDEFGHIJKL',[7,38,7,13,13,22,13,22,48,6,9,8]):
        ws2.column_dimensions[col].width=w

    row = 1
    for ti, tribo in enumerate(TRIBOS):
        grp   = df[df['TRIBO']==tribo].copy()
        cor   = CORES[ti]; cor_l = CORES_LIGHT[ti]
        cor_nm= CORES_NOME[ti]; txt = TEXTO_COR[ti]
        f_c   = (grp['Sexo']=='F').sum(); m_c=(grp['Sexo']=='M').sum()
        aj    = anjos.get(tribo, ANJOS_PADRAO[tribo])

        # Row 1: título da tribo (altura 28)
        ws2.row_dimensions[row].height=28
        c=ws2.cell(row=row,column=1,value=f'Cor: {cor_nm}')
        c.font=Font(bold=True,italic=True,size=13,name='Arial',color=txt)
        c.fill=P(cor); c.alignment=A('left','center'); c.border=BM()
        ws2.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
        c=ws2.cell(row=row,column=5,value=f'Tribo {tribo}')
        c.font=Font(bold=True,italic=True,size=16,name='Arial',color=txt)
        c.fill=P(cor); c.alignment=A('center','center')
        ws2.merge_cells(start_row=row,start_column=5,end_row=row,end_column=12)
        row+=3  # 2 linhas em branco

        # Anjos / Guardião labels
        ws2.cell(row=row,column=1,value='Anjos:').font=F(bold=True)
        ws2.cell(row=row,column=7,value='Anjo Guardião:').font=F(bold=True)
        row+=1

        # Anjo 1 + Guardião (mesma linha)
        c=ws2.cell(row=row,column=1,value=aj['a1'])
        c.font=F(bold=True,size=11); c.fill=P(cor); c.border=BM()
        ws2.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4) # A-D mergeado? não, só até C na foto
        c=ws2.cell(row=row,column=7,value=aj['g'])
        c.font=F(bold=True,size=11,color=txt if txt!='000000' else '000000')
        c.fill=P(cor); c.alignment=A('left')
        ws2.merge_cells(start_row=row,start_column=7,end_row=row,end_column=10)
        row+=1

        # Anjo 2
        c=ws2.cell(row=row,column=1,value=aj['a2'])
        c.font=F(bold=True,size=11); c.fill=P(cor); c.border=BM()
        row+=2  # 1 linha em branco

        # Cabeçalho das colunas
        hdr_row = row
        hdr_vals=['Ficha ✏️','Nome','Peso','Alt','Idade','Cidade','Camiseta','Celular',
                  '⚠️ Alerta / Observações','Sexo',f_c,m_c]
        for j,v in enumerate(hdr_vals,1):
            c=ws2.cell(row=row,column=j,value=v)
            c.font=F(bold=True,color=txt if j not in [11,12] else txt)
            c.fill=P(cor); c.alignment=A('center'); c.border=B('thin') if j>1 else BM()
        row+=1

        # Linhas de campistas: mulheres primeiro, depois homens
        grp_s = pd.concat([grp[grp['Sexo']=='F'].sort_values('Nome'),
                           grp[grp['Sexo']=='M'].sort_values('Nome')])
        data_rows=[]
        for ii,(_,r) in enumerate(grp_s.iterrows()):
            ic = int(r['FC']) in partner
            # Alternância: casal=amarelo, par=branco, ímpar=cor_light
            rf = P('FFF0CC') if ic else (P('FFFFFF') if ii%2==0 else P(cor_l))

            # Col A — FC (editável, azul escuro bold)
            c=ws2.cell(row=row,column=1,value=int(r['FC']))
            c.font=Font(bold=True,size=11,name='Arial',color='1F4E79')
            c.fill=rf; c.border=BM(); c.alignment=A('center')

            # Cols B-J fórmulas
            for col,ci_key,h in [
                (2,'nome','left'),(3,'peso','center'),(4,'altura','center'),
                (5,'idade','center'),(6,'cidade','left'),(7,'camiseta','center'),
                (8,'celular','left'),(10,'sexo','center')]:
                c=ws2.cell(row=row,column=col,value=vl(f'A{row}',CI[ci_key]))
                c.font=F(color='000000'); c.fill=rf
                c.border=B('thin'); c.alignment=A(h)

            # Cols K,L vazias
            for col in [11,12]:
                c=ws2.cell(row=row,column=col); c.fill=rf; c.border=B('thin')

            ws2.row_dimensions[row].height=18
            data_rows.append(row); row+=1

        # Col I — alertas (escrito depois de conhecer o range)
        fc_range=f'$A${data_rows[0]}:$A${data_rows[-1]}'
        for dr_idx, dr in enumerate(data_rows):
            row_fc = ws2.cell(row=dr,column=1).value
            try: row_fc = int(row_fc)
            except: row_fc = 0
            ic_dr = row_fc in partner
            rf_dr = P('FFF0CC') if ic_dr else (P('FFFFFF') if dr_idx%2==0 else P(cor_l))
            c=ws2.cell(row=dr,column=9,value=alert_formula(f'A{dr}',fc_range))
            c.font=F(color='8B4513'); c.fill=rf_dr
            c.border=B('thin'); c.alignment=A('left',wrap=True)

        # Formatação condicional col A
        a_rng=f'A{data_rows[0]}:A{data_rows[-1]}'
        cf_expr(ws2,a_rng,f'ISERROR(VLOOKUP(A{data_rows[0]},{REF},1,0))','C00000','FFFFFF',1)
        cf_expr(ws2,a_rng,
                f'AND(IFERROR(VLOOKUP(A{data_rows[0]},{REF},{CI["parceiro"]},0),0)>0,'
                f'COUNTIF({fc_range},IFERROR(VLOOKUP(A{data_rows[0]},{REF},{CI["parceiro"]},0),0))>0)',
                'FF6600','FFFFFF',2)

        # K/L cabeçalho: COUNTIF + verde/vermelho
        j_range=f'J{data_rows[0]}:J{data_rows[-1]}'
        for col,sexo in [(11,'F'),(12,'M')]:
            c=ws2.cell(row=hdr_row,column=col,value=f'=COUNTIF({j_range},"{sexo}")')
            c.font=F(bold=True,color=txt); c.fill=P(cor)
            c.alignment=A('center'); c.border=B('thin')
            cell_ref=f'{get_column_letter(col)}{hdr_row}'
            cf_expr(ws2,cell_ref,f'{cell_ref}=8','00B050','FFFFFF',2)
            cf_expr(ws2,cell_ref,f'{cell_ref}<>8','C00000','FFFFFF',1)

        # MÉDIA — formato exato da planilha editada
        media_row=row
        c_rng=f'C{data_rows[0]}:C{data_rows[-1]}'
        e_rng=f'E{data_rows[0]}:E{data_rows[-1]}'
        for j in range(1,13):
            val=None
            if j==1: val='MÉDIA>>>>>>>>>'
            elif j==3: val=f'=IFERROR(ROUND(AVERAGE({c_rng}),1),"")'
            elif j==5: val=f'=IFERROR(ROUND(AVERAGE({e_rng}),1),"")'
            c=ws2.cell(row=row,column=j,value=val)
            c.font=F(bold=True); c.fill=P('F2F2F2')
            c.border=BM() if j==1 else (B('thin') if j<=10 else BNONE())
            c.alignment=A('left')
        # Alerta peso/idade: vermelho se desvio >5 da média global
        for cell_ref, g_val in [(f'C{row}',global_peso),(f'E{row}',global_idade)]:
            cf_expr(ws2,cell_ref,f'ABS({cell_ref}-{g_val})>5','C00000','FFFFFF',1)
            cf_expr(ws2,cell_ref,f'ABS({cell_ref}-{g_val})<=5','00B050','FFFFFF',2)
        row+=1

        # Sequências das dinâmicas — col I, exatamente como na planilha
        c=ws2.cell(row=row,column=9,value='Sequências das dinâmicas:')
        c.font=F(bold=True,size=10); c.alignment=A('center')
        ws2.merge_cells(start_row=row,start_column=9,end_row=row,end_column=10)
        row+=1
        for si,d in enumerate([DINAMICAS[(i+ti)%9] for i in range(9)],1):
            c=ws2.cell(row=row,column=9,value=f'{si} - {d}')
            c.font=F(size=10); c.alignment=A('left')
            row+=1
        row+=1  # linha em branco entre tribos

    # ════════════════════════════
    # ABA 3 — Relatorio Chamada Onibus
    # VLOOKUP-driven: col A = FC (editável), B-F = fórmulas automáticas
    # ════════════════════════════
    ws3 = wb.create_sheet('Relatorio Chamada Onibus')
    df_ord = df.sort_values('FC').reset_index(drop=True)

    # Instrução no topo
    inst3 = ws3.cell(row=1, column=1,
        value='✏️  Para trocar campista: edite apenas o número na coluna FC. Nome, cidade, tribo e celular atualizam automaticamente.')
    inst3.font = Font(bold=True, size=10, color='FFFFFF', name='Arial')
    inst3.fill = P('1F4E79')
    inst3.alignment = A('left', wrap=True)
    ws3.merge_cells('A1:F1')
    ws3.row_dimensions[1].height = 28

    row = 2
    for bus in range(4):
        grp = df_ord.iloc[bus*36:(bus+1)*36]
        fc_list = list(grp['FC'])

        # Cabeçalho do ônibus
        c = ws3.cell(row=row, column=1, value=f'onibus 0{bus+1}')
        c.font = F(bold=True, color='FFFFFF', size=11); c.fill = P('1F4E79')
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

        # Cabeçalho colunas
        for j, h in enumerate(['FC. ✏️','CAMPISTAS','CIDADE','TRIBO','FONE:','CELULAR'], 1):
            c = ws3.cell(row=row, column=j, value=h)
            c.font = F(bold=True, color='FFFFFF'); c.fill = P('2F5496')
            c.alignment = A('center'); c.border = B()
        row += 1

        data_start = row
        for ii, fc_val in enumerate(fc_list):
            rf = P('DAEEF3') if ii % 2 == 0 else P('FFFFFF')

            # Col A — FC (editável, azul escuro)
            ca = ws3.cell(row=row, column=1, value=int(fc_val))
            ca.font = Font(bold=True, size=11, name='Arial', color='1F4E79')
            ca.fill = rf; ca.border = B(); ca.alignment = A('center')

            # Col B — Nome via VLOOKUP
            c = ws3.cell(row=row, column=2,
                value=f"=IFERROR(VLOOKUP(A{row},{REF},{CI['nome']},0),"")")
            c.fill = rf; c.border = B(); c.font = F(); c.alignment = A('left')

            # Col C — Cidade via VLOOKUP
            c = ws3.cell(row=row, column=3,
                value=f"=IFERROR(VLOOKUP(A{row},{REF},{CI['cidade']},0),"")")
            c.fill = rf; c.border = B(); c.font = F(); c.alignment = A('left')

            # Col D — Tribo via VLOOKUP (com cor dinâmica via CF)
            c = ws3.cell(row=row, column=4,
                value=f"=IFERROR(VLOOKUP(A{row},{REF},{CI['tribo']},0),"")")
            c.fill = rf; c.border = B(); c.font = F(bold=True); c.alignment = A('center')

            # Col E — vazia (FONE fixo)
            c = ws3.cell(row=row, column=5, value='')
            c.fill = rf; c.border = B(); c.font = F()

            # Col F — Celular via VLOOKUP
            c = ws3.cell(row=row, column=6,
                value=f"=IFERROR(VLOOKUP(A{row},{REF},{CI['celular']},0),"")")
            c.fill = rf; c.border = B(); c.font = F(); c.alignment = A('left')

            # CF col D: colorir tribo automaticamente
            for ti, t in enumerate(TRIBOS):
                ds = DifferentialStyle(
                    fill=PatternFill(start_color=CORES[ti], end_color=CORES[ti], fill_type='solid'),
                    font=Font(bold=True, color=TEXTO_COR[ti], name='Arial'))
                ws3.conditional_formatting.add(f'D{row}',
                    Rule(type='expression', dxf=ds, priority=ti+1,
                         formula=[f'D{row}="{t}"']))

            # CF col A: FC inválido → vermelho
            ds_err = DifferentialStyle(
                fill=PatternFill(start_color='C00000', end_color='C00000', fill_type='solid'),
                font=Font(bold=True, color='FFFFFF', name='Arial'))
            ws3.conditional_formatting.add(f'A{row}',
                Rule(type='expression', dxf=ds_err, priority=10,
                     formula=[f'ISERROR(VLOOKUP(A{row},{REF},1,0))']))

            row += 1
        row += 1  # linha em branco entre ônibus

    for j, w in enumerate([8, 42, 22, 14, 10, 22], 1):
        ws3.column_dimensions[get_column_letter(j)].width = w
    ws3.freeze_panes = 'A3' 

    # ════════════════════════════
    # ABA 4 — CONFLITO CONHECIDOS
    # VLOOKUP-driven: col FC = editável, Nome/Tribo = fórmulas automáticas
    # ════════════════════════════
    ws4 = wb.create_sheet('CONFLITO CONHECIDOS')

    # Instrução
    inst4 = ws4.cell(row=1, column=1,
        value='✏️  Para trocar campista: edite apenas as colunas FC e FC CÔNJUGE. Nome e tribo atualizam automaticamente.')
    inst4.font = Font(bold=True, size=10, color='FFFFFF', name='Arial')
    inst4.fill = P('1F4E79')
    inst4.alignment = A('left', wrap=True)
    ws4.merge_cells('A1:G1')
    ws4.row_dimensions[1].height = 28

    # ── Seção 1: Casais ──
    c = ws4.cell(row=2, column=1, value='CASAIS – ACAMPAMENTO ESPÍRITO EMPREENDEDOR')
    c.font = F(bold=True, color='FFFFFF', size=12); c.fill = P('1F4E79')
    ws4.merge_cells('A2:G2')

    for j, h in enumerate(['#', 'FC ✏️', 'NOME', 'TRIBO', 'FC CÔNJUGE ✏️', 'NOME CÔNJUGE', 'TRIBO CÔNJUGE'], 1):
        c = ws4.cell(row=3, column=j, value=h)
        c.font = F(bold=True, color='FFFFFF'); c.fill = P('C00000')
        c.alignment = A('center'); c.border = B()

    seen = set(); row = 4; num = 1
    for fc_a, fc_b in [(a, b) for a, b in partner.items() if a < b]:
        if fc_a in seen: continue
        seen.add(fc_a); seen.add(fc_b)
        rows_a = df[df['FC'] == fc_a]; rows_b = df[df['FC'] == fc_b]
        if rows_a.empty or rows_b.empty: continue
        tia = TRIBOS.index(rows_a.iloc[0]['TRIBO'])
        tib = TRIBOS.index(rows_b.iloc[0]['TRIBO'])
        rf  = P('FFF0CC') if num % 2 == 0 else P('FFFFFF')

        # Col A: número sequencial
        c = ws4.cell(row=row, column=1, value=num)
        c.fill = rf; c.font = F(); c.alignment = A('center'); c.border = B()

        # Col B: FC campista 1 (editável)
        c = ws4.cell(row=row, column=2, value=int(fc_a))
        c.fill = rf; c.font = Font(bold=True, size=11, name='Arial', color='1F4E79')
        c.alignment = A('center'); c.border = B()

        # Col C: Nome campista 1 via VLOOKUP
        c = ws4.cell(row=row, column=3,
            value=f"=IFERROR(VLOOKUP(B{row},{REF},{CI['nome']},0),"")")
        c.fill = rf; c.font = F(); c.alignment = A('left'); c.border = B()

        # Col D: Tribo campista 1 via VLOOKUP
        c = ws4.cell(row=row, column=4,
            value=f"=IFERROR(VLOOKUP(B{row},{REF},{CI['tribo']},0),"")")
        c.fill = rf; c.font = F(bold=True); c.alignment = A('center'); c.border = B()

        # Col E: FC cônjuge (editável)
        c = ws4.cell(row=row, column=5, value=int(fc_b))
        c.fill = rf; c.font = Font(bold=True, size=11, name='Arial', color='1F4E79')
        c.alignment = A('center'); c.border = B()

        # Col F: Nome cônjuge via VLOOKUP
        c = ws4.cell(row=row, column=6,
            value=f"=IFERROR(VLOOKUP(E{row},{REF},{CI['nome']},0),"")")
        c.fill = rf; c.font = F(); c.alignment = A('left'); c.border = B()

        # Col G: Tribo cônjuge via VLOOKUP
        c = ws4.cell(row=row, column=7,
            value=f"=IFERROR(VLOOKUP(E{row},{REF},{CI['tribo']},0),"")")
        c.fill = rf; c.font = F(bold=True); c.alignment = A('center'); c.border = B()

        # CF: colorir tribos via VLOOKUP nas colunas D e G
        for col_tribo in [4, 7]:
            for ti, t in enumerate(TRIBOS):
                fc_col = 'B' if col_tribo == 4 else 'E'
                ds = DifferentialStyle(
                    fill=PatternFill(start_color=CORES[ti], end_color=CORES[ti], fill_type='solid'),
                    font=Font(bold=True, color=TEXTO_COR[ti], name='Arial'))
                ws4.conditional_formatting.add(f'{get_column_letter(col_tribo)}{row}',
                    Rule(type='expression', dxf=ds, priority=ti+1,
                         formula=[f'{get_column_letter(col_tribo)}{row}="{t}"']))

        # CF: FC inválido → vermelho nas cols B e E
        ds_err = DifferentialStyle(
            fill=PatternFill(start_color='C00000', end_color='C00000', fill_type='solid'),
            font=Font(bold=True, color='FFFFFF', name='Arial'))
        for fc_col in ['B', 'E']:
            ws4.conditional_formatting.add(f'{fc_col}{row}',
                Rule(type='expression', dxf=ds_err, priority=10,
                     formula=[f'ISERROR(VLOOKUP({fc_col}{row},{REF},1,0))']))

        row += 1; num += 1

    # ── Seção 2: Demais conhecidos ──
    row += 1
    c = ws4.cell(row=row, column=1, value='DEMAIS CAMPISTAS COM CONHECIDOS DECLARADOS:')
    c.font = F(bold=True, size=11); c.fill = P('F2F2F2')
    ws4.merge_cells(f'A{row}:G{row}'); row += 1

    for j, h in enumerate(['FC ✏️', 'NOME', 'TRIBO', 'CONHECIDOS DECLARADOS'], 1):
        c = ws4.cell(row=row, column=j, value=h)
        c.font = F(bold=True, color='FFFFFF'); c.fill = P('404040')
        c.alignment = A('center'); c.border = B()
    row += 1

    for ii, (_, r) in enumerate(df.iterrows()):
        if int(r['FC']) in partner: continue
        conf = str(r['Conflitos']).strip()
        if is_skip(conf): continue
        ti  = TRIBOS.index(r['TRIBO'])
        rf  = P('F2F2F2') if ii % 2 == 0 else P('FFFFFF')

        # Col A: FC (editável)
        c = ws4.cell(row=row, column=1, value=int(r['FC']))
        c.fill = rf; c.font = Font(bold=True, size=11, name='Arial', color='1F4E79')
        c.alignment = A('center'); c.border = B()

        # Col B: Nome via VLOOKUP
        c = ws4.cell(row=row, column=2,
            value=f"=IFERROR(VLOOKUP(A{row},{REF},{CI['nome']},0),"")")
        c.fill = rf; c.font = F(); c.alignment = A('left'); c.border = B()

        # Col C: Tribo via VLOOKUP
        c = ws4.cell(row=row, column=3,
            value=f"=IFERROR(VLOOKUP(A{row},{REF},{CI['tribo']},0),"")")
        c.fill = rf; c.font = F(bold=True); c.alignment = A('center'); c.border = B()

        # Col D: Conflitos (texto estático — não muda com o FC)
        c = ws4.cell(row=row, column=4, value=conf)
        c.fill = rf; c.font = F(); c.alignment = A('left', wrap=True); c.border = B()

        # CF: cor da tribo col C
        for ti2, t in enumerate(TRIBOS):
            ds = DifferentialStyle(
                fill=PatternFill(start_color=CORES[ti2], end_color=CORES[ti2], fill_type='solid'),
                font=Font(bold=True, color=TEXTO_COR[ti2], name='Arial'))
            ws4.conditional_formatting.add(f'C{row}',
                Rule(type='expression', dxf=ds, priority=ti2+1,
                     formula=[f'C{row}="{t}"']))

        # CF: FC inválido → vermelho col A
        ds_err = DifferentialStyle(
            fill=PatternFill(start_color='C00000', end_color='C00000', fill_type='solid'),
            font=Font(bold=True, color='FFFFFF', name='Arial'))
        ws4.conditional_formatting.add(f'A{row}',
            Rule(type='expression', dxf=ds_err, priority=10,
                 formula=[f'ISERROR(VLOOKUP(A{row},{REF},1,0))']))

        row += 1

    for j, w in enumerate([8, 38, 14, 55, 12, 38, 14], 1):
        ws4.column_dimensions[get_column_letter(j)].width = w
    ws4.freeze_panes = 'A4' 

    wb.save(output_path)
    return output_path

# ─────────────────────────────────────────────
# 5. TESTES
# ─────────────────────────────────────────────
def rodar_testes(df, partner):
    erros=[]
    avisos=[]

    # Teste 1: total 144 campistas
    if len(df) != 144:
        avisos.append(f"Total de campistas: {len(df)} (esperado 144)")

    # Teste 2: distribuição por tribo — aviso se M≠F (pode ter inscrições desbalanceadas)
    for t in TRIBOS:
        g=df[df['TRIBO']==t]
        m=(g['Sexo']=='M').sum(); f=(g['Sexo']=='F').sum()
        if m != f:
            avisos.append(f"Tribo {t}: {f}F {m}M (inscrições desbalanceadas — distribui o melhor possível)")

    # Teste 3: sem casal na mesma tribo
    for fc_a,fc_b in [(a,b) for a,b in partner.items() if a<b]:
        ra=df[df['FC']==fc_a]; rb=df[df['FC']==fc_b]
        if ra.empty or rb.empty: continue
        ta=ra.iloc[0]['TRIBO']; tb=rb.iloc[0]['TRIBO']
        if ta==tb:
            na=ra.iloc[0]['Nome']; nb=rb.iloc[0]['Nome']
            erros.append(f"CASAL NA MESMA TRIBO: FC{fc_a} {na} e FC{fc_b} {nb} → {ta}")

    # Teste 4: balanceamento peso/idade
    peso_med  = [df[df['TRIBO']==t]['Peso_N'].mean() for t in TRIBOS]
    idade_med = [df[df['TRIBO']==t]['Idade_N'].mean() for t in TRIBOS]
    dp_peso   = np.std(peso_med); dp_idade = np.std(idade_med)
    if dp_peso > 5:
        avisos.append(f"Desv.Pad. peso entre tribos: {dp_peso:.1f}kg (aceitável ≤5)")
    if dp_idade > 5:
        avisos.append(f"Desv.Pad. idade entre tribos: {dp_idade:.1f}a (aceitável ≤5)")

    return erros, avisos

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main(input_path, output_path=None, anjos=None, inscricoes_path=None):
    if output_path is None:
        base = input_path.rsplit('.',1)[0]
        output_path = base + '_TRIBOS_GERADO.xlsx'

    print(f"\n📂 Lendo arquivo: {input_path}")

    # Detecta automaticamente o tipo de arquivo
    # Se tem coluna 'Cupom' ou 'Categoria' = planilha de inscrições original
    # Se tem aba 'Cadastro Geral Campistas' = planilha base já processada
    is_inscricoes = False
    try:
        wb_check = load_workbook(input_path, read_only=True)
        sheet_names = wb_check.sheetnames
        if 'Cadastro Geral Campistas' in sheet_names:
            is_inscricoes = False
            print("   Tipo detectado: Planilha base (aba Cadastro Geral Campistas)")
        else:
            is_inscricoes = True
            print("   Tipo detectado: Planilha de inscrições original")
        wb_check.close()
    except Exception:
        is_inscricoes = True

    if is_inscricoes:
        df = ler_inscricoes(input_path)
    else:
        df = ler_base(input_path)
        print(f"   {len(df)} campistas | M:{(df['Sexo']=='M').sum()} F:{(df['Sexo']=='F').sum()}")

    # Planilha de inscrições separada (opcional, para modo base)
    df_insc = None
    if inscricoes_path:
        try:
            df_insc = pd.read_excel(inscricoes_path, header=0)
            print(f"   Planilha de inscrições extra carregada: {len(df_insc)} registros")
        except Exception as e:
            print(f"   Aviso: não foi possível ler inscrições: {e}")

    print("\n🔍 Identificando casais (Cupom CASAL + Categoria + Parceiro FC + Conflitos)...")
    partner = identificar_casais(df, df_inscricoes=df_insc)
    print(f"   {len(partner)//2} casais identificados")

    print("\n⚖️  Balanceando tribos (snake draft peso+idade)...")
    df = atribuir_tribos(df, partner)
    for t in TRIBOS:
        g=df[df['TRIBO']==t]
        m=(g['Sexo']=='M').sum(); f=(g['Sexo']=='F').sum()
        print(f"   {t:10s}: M={m} F={f}  Peso={g['Peso_N'].mean():.1f}  Idade={g['Idade_N'].mean():.1f}")

    print("\n✅ Rodando testes...")
    erros, avisos = rodar_testes(df, partner)
    if erros:
        print("  ❌ ERROS:")
        for e in erros: print(f"     {e}")
    else:
        print("  ✅ Todos os testes passaram!")
    if avisos:
        print("  ⚠️  Avisos:")
        for a in avisos: print(f"     {a}")

    print(f"\n📝 Gerando planilha: {output_path}")
    gerar(df, partner, anjos=anjos, output_path=output_path)
    print(f"✅ Pronto!\n")
    return output_path


# ─────────────────────────────────────────────
# ATUALIZAÇÃO — compara nova lista com planilha
# já gerada, remove cancelados, insere novos
# ─────────────────────────────────────────────
def atualizar(planilha_gerada_path, nova_inscricoes_path, output_path=None, anjos=None):
    if output_path is None:
        base = planilha_gerada_path.rsplit('.', 1)[0]
        output_path = base + '_ATUALIZADO.xlsx'

    print(f"\n📂 Planilha anterior: {planilha_gerada_path}")
    df_atual = ler_base(planilha_gerada_path)
    print(f"   {len(df_atual)} campistas na versao anterior")

    print(f"\n📂 Nova lista de inscricoes: {nova_inscricoes_path}")
    df_nova = ler_inscricoes(nova_inscricoes_path)
    print(f"   {len(df_nova)} inscricoes ativas na nova lista")

    def norm(s):
        return ' '.join(str(s).lower().strip().split())

    nomes_novos  = {norm(r['Nome']): r for _, r in df_nova.iterrows()}
    nomes_atuais = {norm(r['Nome']): r for _, r in df_atual.iterrows()}

    # Quem saiu
    saiu = [n for n in nomes_atuais if n not in nomes_novos]
    print(f"\n🗑️  Removidos (desistencias/cancelamentos): {len(saiu)}")
    for n in saiu:
        r = nomes_atuais[n]
        tribo = r.get('TRIBO', '?') if hasattr(r, 'get') else '?'
        print(f"   FC{int(r['FC']):3d}  {r['Nome']}  (Tribo: {tribo})")

    # Quem entrou
    entrou = [n for n in nomes_novos if n not in nomes_atuais]
    print(f"\n✨ Novos campistas: {len(entrou)}")
    for n in sorted(entrou):
        r = nomes_novos[n]
        print(f"   {r['Nome']}  ({r['Sexo']})")

    # Base mantida — atualiza dados que podem ter mudado
    df_mantidos = df_atual[df_atual['Nome'].apply(norm).isin(nomes_novos)].copy()
    for idx, row in df_mantidos.iterrows():
        n = norm(row['Nome'])
        novo = nomes_novos.get(n, {})
        for campo in ['Camiseta','Peso','Altura','Celular','Cidade',
                      'Conflitos','C1Nome','C1Tel','C2Nome','C2Tel','Alergias','Nasc']:
            v = novo.get(campo) if hasattr(novo, 'get') else None
            if v not in (None, '', 'nan'):
                df_mantidos.at[idx, campo] = v

    # Atribui FC aos novos continuando do maior existente
    max_fc = int(df_mantidos['FC'].max()) if len(df_mantidos) else 0
    novos_rows = []
    for i, n in enumerate(sorted(entrou), 1):
        r = dict(nomes_novos[n])
        r['FC']      = max_fc + i
        r['TRIBO']   = ''
        r['ParcFC']  = ''
        r['Peso_N']  = _clean_peso(r.get('Peso'))
        r['Idade_N'] = r.get('Idade')
        r['Alt_N']   = _clean_altura(r.get('Altura'))
        # Garante campos auxiliares
        if '_cupom' not in r: r['_cupom'] = ''
        if '_categ' not in r: r['_categ'] = ''
        if '_lider' not in r: r['_lider'] = ''
        novos_rows.append(r)

    df_novos = pd.DataFrame(novos_rows) if novos_rows else pd.DataFrame(columns=df_mantidos.columns)
    for col in ['Peso_N','Idade_N','Alt_N','_cupom','_categ','_lider']:
        if col not in df_novos.columns:
            df_novos[col] = '' if col.startswith('_') else None

    df_total = pd.concat([df_mantidos, df_novos], ignore_index=True)
    df_total['Peso_N']  = pd.to_numeric(df_total['Peso_N'],  errors='coerce')
    df_total['Idade_N'] = pd.to_numeric(df_total['Idade_N'], errors='coerce')
    df_total['Alt_N']   = pd.to_numeric(df_total['Alt_N'],   errors='coerce')

    print("\n🔍 Identificando casais...")
    partner = identificar_casais(df_total)
    print(f"   {len(partner)//2} casais")

    # Distribui novos nas tribos com menor contagem do sexo
    if novos_rows:
        print(f"\n⚖️  Distribuindo {len(novos_rows)} novo(s)...")
        contagem = {t: {'M': 0, 'F': 0} for t in TRIBOS}
        for _, r in df_mantidos.iterrows():
            t = r.get('TRIBO', '') if hasattr(r, 'get') else str(r['TRIBO']) if 'TRIBO' in r.index else ''
            if t in contagem:
                contagem[t][r['Sexo']] += 1

        for _, row in df_novos.iterrows():
            fc   = int(row['FC'])
            sexo = row['Sexo']
            p_fc = partner.get(fc)
            for t in sorted(TRIBOS, key=lambda x: contagem[x][sexo]):
                p_tribo = None
                if p_fc:
                    rows_p = df_total[df_total['FC'] == p_fc]
                    p_tribo = rows_p.iloc[0]['TRIBO'] if not rows_p.empty else None
                if p_tribo != t:
                    df_total.loc[df_total['FC'] == fc, 'TRIBO'] = t
                    contagem[t][sexo] += 1
                    print(f"   FC{fc:3d} {row['Nome'][:30]} → {t}")
                    break

    # Status final das tribos
    print("\n📊 Contagem final por tribo:")
    for t in TRIBOS:
        g = df_total[df_total['TRIBO'] == t]
        m = (g['Sexo'] == 'M').sum(); f = (g['Sexo'] == 'F').sum()
        st = "✅" if m == 8 and f == 8 else f"⚠️  ATENÇÃO"
        print(f"   {t:10s}: M={m} F={f}  {st}")

    print("\n✅ Rodando testes...")
    erros, avisos = rodar_testes(df_total, partner)
    if erros:
        print("  ❌ ERROS:"); [print(f"     {e}") for e in erros]
    else:
        print("  ✅ Todos os testes passaram!")
    if avisos:
        [print(f"  ⚠️  {a}") for a in avisos]

    print(f"\n📝 Gerando: {output_path}")
    gerar(df_total, partner, anjos=anjos, output_path=output_path)
    print(f"✅ Concluido! {len(df_total)} campistas · {len(saiu)} removidos · {len(entrou)} adicionados\n")
    return output_path


if __name__ == '__main__':
    import sys, os
    if len(sys.argv) < 2:
        print("Uso:")
        print("  Gerar (1ª vez):")
        print("    python3 gerar_tribos.py <inscricoes.csv|xlsx> [saida.xlsx]")
        print()
        print("  Atualizar (desistências/novos):")
        print("    python3 gerar_tribos.py --atualizar <planilha_gerada.xlsx> <nova_inscricoes.csv|xlsx> [saida.xlsx]")
        sys.exit(1)

    if sys.argv[1] == '--atualizar':
        if len(sys.argv) < 4:
            print("Uso: python3 gerar_tribos.py --atualizar <planilha_gerada.xlsx> <nova_inscricoes.csv|xlsx> [saida.xlsx]")
            sys.exit(1)
        gerada = sys.argv[2]
        nova   = sys.argv[3]
        saida  = sys.argv[4] if len(sys.argv) > 4 else None
        atualizar(gerada, nova, output_path=saida)
    else:
        inp  = sys.argv[1]
        out  = sys.argv[2] if len(sys.argv) > 2 else None
        main(inp, out)
