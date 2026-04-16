"""
atualizar_tribos.py
===================
Atualiza a planilha preservando FCs e tribos de quem já está.

REGRAS ABSOLUTAS:
  1. Quem continua → mesmo FC, mesma tribo, mesmo lugar. NUNCA move.
  2. Cancelado      → FC fica vazio (disponível para reutilizar)
  3. Novato         → recebe o FC do cancelado (ou próximo número livre)
  4. NUNCA renumera quem já está na planilha

Uso:
  python3 atualizar_tribos.py novas_inscricoes.csv

Precisa do VIII_MODELO.xlsx na mesma pasta.
"""

import sys, os, copy
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

MODELO      = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'VIII_MODELO.xlsx')
TRIBOS      = ['Simeão','Rubem','Judá','Levi','Benjamim','Issacar','Gad','Zabulom','EFRAIM']
TRIBO_STARTS= [1,36,71,106,141,176,211,246,281]
TRIBO_COR   = {'Simeão':'FFC000','Rubem':'FF0000','Judá':'FF6600','Levi':'FF99CC',
               'Benjamim':'0070C0','Issacar':'00B050','Gad':'8B4513','Zabulom':'000000','EFRAIM':'BFBFBF'}
TRIBO_TXT   = {'Simeão':'000000','Rubem':'FFFFFF','Judá':'FFFFFF','Levi':'000000',
               'Benjamim':'FFFFFF','Issacar':'FFFFFF','Gad':'FFFFFF','Zabulom':'FFFFFF','EFRAIM':'000000'}
TRIBO_LIGHT = {'Simeão':'FFF2CC','Rubem':'FFE8E8','Judá':'FFE5CC','Levi':'FFD6EC',
               'Benjamim':'DAEEF3','Issacar':'EAF6E2','Gad':'F5E6DA','Zabulom':'333333','EFRAIM':'F5F5F5'}

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from gerar_tribos import ler_inscricoes, identificar_casais

# ── Helpers ──────────────────────────────────────────────────────
def sv(v):
    s = str(v).strip() if v is not None else ''
    return '' if s.lower() in ['nan','none','nat'] else s

def tb(left='thin'):
    return Border(left=Side(style=left), right=Side(style='thin'),
                  top=Side(style='thin'),  bottom=Side(style='thin'))

def safe_write(ws, row, col, value, **kwargs):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value
    for k, v in kwargs.items():
        setattr(cell, k, v)

def build_merge_map(ws):
    m = set()
    for mc in ws.merged_cells.ranges:
        for r in range(mc.min_row, mc.max_row+1):
            for c in range(mc.min_col, mc.max_col+1):
                if not (r==mc.min_row and c==mc.min_col):
                    m.add((r,c))
    return m

def normalizar(arr):
    mn, mx = np.nanmin(arr), np.nanmax(arr)
    if mx == mn: return np.zeros_like(arr, dtype=float)
    return (arr - mn) / (mx - mn)

# ── Lê estado atual do modelo ─────────────────────────────────────
def ler_modelo(wb):
    """
    Lê o Cadastro Geral e retorna:
      - dict fc → {nome, tribo, sexo, ...}   (mapa do estado atual)
      - dict nome_lower → fc                 (para reconhecer continuantes)
    """
    ws = wb['Cadastro Geral Campistas']
    hdr = {cell.value: j for j, cell in enumerate(ws[1], 1) if cell.value}
    col = lambda k, d: hdr.get(k, d)
    cfc = col('FICHA',1); cnome=col('NOME',2); ctribo=col('TRIBO',17)
    csexo=col('SEXO',4);  cnasc=col('DATA DE NASCIMENTO',16)

    fc_map   = {}   # fc → dados
    nome_map = {}   # nome_lower → fc
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[cfc-1]: continue
        try: fc = int(row[cfc-1])
        except: continue
        nome  = sv(row[cnome-1])
        tribo = sv(row[ctribo-1])
        sexo  = sv(row[csexo-1])
        if not nome: continue
        fc_map[fc]             = {'nome':nome,'tribo':tribo,'sexo':sexo}
        nome_map[nome.lower()] = fc
    return fc_map, nome_map

# ── Distribui novatos nas vagas abertas ───────────────────────────
def distribuir_novatos(novatos, ocupacao, partner_fc):
    """
    novatos    : lista de dicts com dados do campista (sem FC ainda)
    ocupacao   : dict {tribo: {M: n, F: n}} — quantos já tem por sexo
    partner_fc : dict fc → fc_parceiro (casais já com FCs definitivos)
    Retorna lista de dicts com campo TRIBO preenchido.
    """
    # Vagas disponíveis
    vagas = {t: {'M': 8 - ocupacao.get(t,{}).get('M',0),
                 'F': 8 - ocupacao.get(t,{}).get('F',0)}
             for t in TRIBOS}

    # Score para balancear
    pesos  = [float(n.get('Peso') or 0)  for n in novatos]
    idades = [float(n.get('Idade') or 0) for n in novatos]
    arr_p  = normalizar(np.array(pesos,  dtype=float))
    arr_i  = normalizar(np.array(idades, dtype=float))
    for i,n in enumerate(novatos):
        n['_sc'] = arr_p[i]*0.5 + arr_i[i]*0.5

    novatos_s = sorted(novatos, key=lambda x: x['_sc'])
    assigns   = {}

    for nov in novatos_s:
        fc   = nov['FC']
        sexo = nov.get('Sexo','M')[0].upper()
        p_fc = partner_fc.get(fc)
        p_t  = assigns.get(p_fc)

        melhor, melhor_v = None, -1
        for t in TRIBOS:
            v = vagas[t][sexo]
            if v > 0 and p_t != t and v > melhor_v:
                melhor, melhor_v = t, v

        if not melhor:   # sem vaga ideal → menos cheio
            melhor = sorted(TRIBOS, key=lambda t: vagas[t][sexo])[-1]

        assigns[fc]       = melhor
        vagas[melhor][sexo] = max(0, vagas[melhor][sexo]-1)

    for nov in novatos:
        nov['TRIBO'] = assigns.get(nov['FC'], TRIBOS[0])

    return novatos

# ── Testes ────────────────────────────────────────────────────────
def rodar_testes(df, partner):
    erros, avisos = [], []
    total = len(df)
    if total != 144:
        avisos.append(f"Total: {total} campistas (esperado 144)")
    for t in TRIBOS:
        g = df[df['TRIBO']==t]
        m = (g['Sexo']=='M').sum()
        f = (g['Sexo']=='F').sum()
        if m != f:
            avisos.append(f"{t}: {f}F {m}M")
    for a,b in [(a,b) for a,b in partner.items() if a<b]:
        ra = df[df['FC']==a]; rb = df[df['FC']==b]
        if ra.empty or rb.empty: continue
        if ra.iloc[0]['TRIBO'] == rb.iloc[0]['TRIBO']:
            erros.append(f"CASAL NA MESMA TRIBO: FC{a} + FC{b} → {ra.iloc[0]['TRIBO']}")
    return erros, avisos

# ── Atualiza Cadastro Geral ───────────────────────────────────────
def atualizar_cadastro(ws_cad, df_final, partner):
    merged = build_merge_map(ws_cad)
    # Captura estilo de referência da linha 2
    ref = {}
    for c in range(1,20):
        cell = ws_cad.cell(row=2, column=c)
        ref[c] = {'font':copy.copy(cell.font),'fill':copy.copy(cell.fill),
                  'border':copy.copy(cell.border),'alignment':copy.copy(cell.alignment)}

    # Limpa tudo (exceto cabeçalho)
    for r in range(2, ws_cad.max_row+1):
        for c in range(1, 20):
            if (r,c) not in merged:
                ws_cad.cell(row=r, column=c).value = None

    # Garante cabeçalho na coluna 19
    if not ws_cad.cell(row=1, column=19).value:
        ws_cad.cell(row=1, column=19).value = 'DOCUMENTO'

    # Reescreve na ordem do FC
    df_s = df_final.sort_values('FC').reset_index(drop=True)
    for i, (_, r) in enumerate(df_s.iterrows(), 2):
        fc   = int(r['FC'])
        ic   = fc in partner
        parc = partner.get(fc, '')
        is_n = r.get('_status','') == 'novato'

        def to_num(v):
            try:
                s = str(v).replace(',','.').replace('kg','').replace('m','').strip()
                f = float(s)
                return int(f) if f == int(f) else f
            except: return sv(v)

        vals = [fc, r['Nome'], sv(r.get('Camiseta','')), ('M' if str(r.get('Sexo','')).strip().upper().startswith('M') else 'F'),
                to_num(r.get('Peso','')),  to_num(r.get('Altura','')), to_num(r.get('Idade','')),
                sv(r.get('Celular','')), sv(r.get('Cidade','')), sv(r.get('Conflitos','')),
                sv(r.get('C1Nome','')), sv(r.get('C1Tel','')),
                sv(r.get('C2Nome','')), sv(r.get('C2Tel','')),
                sv(r.get('Alergias','')), sv(r.get('Nasc','')),
                r['TRIBO'], parc, sv(r.get('DocNum',''))]

        for j, val in enumerate(vals, 1):
            if (i,j) in merged: continue
            cell = ws_cad.cell(row=i, column=j, value=val)
            s = ref[j]
            cell.font=copy.copy(s['font']); cell.fill=copy.copy(s['fill'])
            cell.border=copy.copy(s['border']); cell.alignment=copy.copy(s['alignment'])

            if j == 17:   # tribo colorida
                cell.fill = PatternFill('solid', fgColor='FF'+TRIBO_COR[r['TRIBO']])
                cell.font = Font(bold=True, name='Arial', size=10,
                                 color='FF'+TRIBO_TXT[r['TRIBO']])
            elif j == 18 and parc:  # parceiro marrom
                cell.font = Font(bold=True, name='Arial', size=10, color='FF8B4513')
            elif ic and j not in [17,18]:
                cell.fill = PatternFill('solid', fgColor='FFFFF0CC')
            elif is_n and not ic and j not in [17,18]:
                cell.fill = PatternFill('solid', fgColor='FFE8F4FF')  # novato = azul claro

    # Corrige idades negativas
    from datetime import datetime as _dt
    hdr = {cell.value: j for j, cell in enumerate(ws_cad[1], 1) if cell.value}
    col_nasc  = hdr.get('DATA DE NASCIMENTO', 16)
    col_idade = hdr.get('IDADE', 7)
    for row in ws_cad.iter_rows(min_row=2):
        nasc_c  = row[col_nasc-1]
        idade_c = row[col_idade-1]
        if isinstance(nasc_c, MergedCell) or isinstance(idade_c, MergedCell): continue
        try:
            if float(str(idade_c.value or 0)) < 0:
                d = pd.to_datetime(str(nasc_c.value or '').strip())
                if d.year > _dt.now().year: d = d.replace(year=d.year-100)
                idade_c.value = int((_dt(2026,4,16)-d).days/365.25)
                print(f"   Idade corrigida: {row[1].value} → {idade_c.value} anos")
        except: pass

# ── Atualiza Relatorio Familia Provas ─────────────────────────────
def atualizar_familia(ws_fp, df_final, partner):
    merged = build_merge_map(ws_fp)
    # Adiciona formatação condicional: conflito ⚠️ → fonte vermelha negrito
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.styles import Color
    red_font  = Font(bold=True, color=Color(rgb='FFCC0000'), name='Arial', size=10)
    red_style = DifferentialStyle(font=red_font)
    for _t, _s in zip(TRIBOS, TRIBO_STARTS):
        _cs = _s + 8; _ce = _s + 23
        _rng = f'I{_cs}:I{_ce}'
        # Remove CF existente nesse range para evitar duplicatas
        for _cf in list(ws_fp.conditional_formatting._cf_rules.keys()):
            if str(_cf) == _rng:
                try: del ws_fp.conditional_formatting._cf_rules[_cf]
                except: pass
        ws_fp.conditional_formatting.add(_rng, Rule(
            type='expression', dxf=red_style, stopIfTrue=True,
            formula=[f'NOT(ISERROR(SEARCH("⚠",I{_cs})))']
        ))

    # Garante fill cinza na linha MÉDIA antes de tudo
    cinza_fill = PatternFill('solid', fgColor='FFF2F2F2')
    for _t, _s in zip(TRIBOS, TRIBO_STARTS):
        _mr = _s + 24
        for _c in range(1, 11):
            _cell = ws_fp.cell(row=_mr, column=_c)
            if not isinstance(_cell, MergedCell):
                _cell.fill = cinza_fill

    for tribo, start in zip(TRIBOS, TRIBO_STARTS):
        grp = df_final[df_final['TRIBO']==tribo].copy()
        grp_s = pd.concat([
            grp[grp['Sexo']=='F'].sort_values('Nome'),
            grp[grp['Sexo']=='M'].sort_values('Nome')
        ]).reset_index(drop=True)

        for slot in range(16):
            row = start + 8 + slot
            if (row,1) in merged: continue
            cell = ws_fp.cell(row=row, column=1)
            if slot < len(grp_s):
                fc  = int(grp_s.iloc[slot]['FC'])
                ic  = fc in partner
                bg  = 'FFFFF0CC' if ic else \
                      ('FFFFFFFF' if slot%2==0 else 'FF'+TRIBO_LIGHT[tribo])
                cell.value = fc
                for col in range(1, 11):
                    if (row,col) in merged: continue
                    ws_fp.cell(row=row,column=col).fill = PatternFill('solid', fgColor=bg)
            else:
                cell.value = ''
                for col in range(1, 11):
                    if (row,col) in merged: continue
                    ws_fp.cell(row=row,column=col).fill = PatternFill('solid', fgColor='FFFFFFFF')

# ── Atualiza Onibus ───────────────────────────────────────────────
def atualizar_onibus(ws_on, df_final):
    merged = build_merge_map(ws_on)
    df_ord = df_final.sort_values('FC').reset_index(drop=True)
    # Descobre onde começam os dados de cada ônibus
    bus_rows = []
    for r in range(1, ws_on.max_row+1):
        val = ws_on.cell(row=r, column=1).value
        try:
            if int(str(val).strip()) == 1: bus_rows.append(r)
        except: pass

    for bus_idx, data_start in enumerate(bus_rows):
        grp = df_ord.iloc[bus_idx*36:(bus_idx+1)*36]
        for ii, (_, r) in enumerate(grp.iterrows()):
            row = data_start + ii
            for col, val in [(1,int(r['FC'])),(2,r['Nome']),(3,sv(r.get('Cidade',''))),(6,sv(r.get('Celular','')))]:
                if (row,col) not in merged:
                    ws_on.cell(row=row, column=col).value = val
            if (row,4) not in merged:
                c4 = ws_on.cell(row=row, column=4)
                c4.value = r['TRIBO']
                c4.fill  = PatternFill('solid', fgColor='FF'+TRIBO_COR[r['TRIBO']])
                c4.font  = Font(bold=True, name='Arial', size=10, color='FF'+TRIBO_TXT[r['TRIBO']])

# ── FUNÇÃO PRINCIPAL ──────────────────────────────────────────────
def atualizar(csv_path, output_path=None):
    if not output_path:
        output_path = csv_path.rsplit('.',1)[0] + '_ATUALIZADO.xlsx'

    if not os.path.exists(MODELO):
        print(f"\n❌  Modelo não encontrado: {MODELO}")
        print("   Coloque VIII_MODELO.xlsx na mesma pasta.")
        sys.exit(1)

    print(f"\n{'='*55}\n  ATUALIZANDO TRIBOS\n{'='*55}")

    # 1. Lê novas inscrições
    print(f"\n📂 Lendo: {os.path.basename(csv_path)}")
    df_novo = ler_inscricoes(csv_path)
    print(f"   {len(df_novo)} ativos no CSV")

    # 2. Lê estado atual do modelo
    print(f"📋 Carregando modelo...")
    wb = load_workbook(MODELO)
    fc_map, nome_map = ler_modelo(wb)
    print(f"   Modelo: {len(fc_map)} campistas com FCs definidos")

    # Mapa FC → dados do modelo (para preservar)
    # nome_map: nome_lower → FC antigo

    # 3. Classifica cada campista novo
    continuantes = []  # já estavam → preserva FC e tribo
    novatos_raw  = []  # entraram agora → recebe FC vago ou próximo

    for _, r in df_novo.iterrows():
        nome_l = r['Nome'].strip().lower()
        if nome_l in nome_map:
            fc_ant = nome_map[nome_l]
            r2 = r.copy()
            r2['FC']      = fc_ant                    # mantém FC original
            r2['TRIBO']   = fc_map[fc_ant]['tribo']   # mantém tribo original
            r2['_status'] = 'continuante'
            continuantes.append(r2)
        else:
            r2 = r.copy()
            r2['_status'] = 'novato'
            novatos_raw.append(r2)

    # 4. Identifica FCs vagos (cancelados)
    fcs_continuantes = {int(r['FC']) for r in continuantes}
    fcs_cancelados   = sorted(set(fc_map.keys()) - fcs_continuantes)
    # FCs para novatos: primeiro os vagos, depois novos números
    max_fc = max(fc_map.keys()) if fc_map else 0
    fila_fcs = fcs_cancelados + list(range(max_fc+1, max_fc+1+len(novatos_raw)+5))

    # Atribui FCs aos novatos
    for i, r in enumerate(novatos_raw):
        r['FC'] = fila_fcs[i]

    print(f"\n📊 Análise:")
    print(f"   ✅ Continuantes (FC e tribo preservados): {len(continuantes)}")
    print(f"   🆕 Novatos:                               {len(novatos_raw)}")
    print(f"   ❌ Cancelados (FCs vagos):                {len(fcs_cancelados)}")
    if fcs_cancelados:
        # Mostra quem cancelou
        fc_to_nome = {fc: data['nome'] for fc, data in fc_map.items()}
        for fc in fcs_cancelados:
            print(f"      FC{fc:3d} {fc_to_nome.get(fc,'?')} → removido")
    if novatos_raw:
        print(f"   Novatos e seus novos FCs:")
        for r in novatos_raw:
            print(f"      FC{int(r['FC']):3d} {r['Nome'].strip()} (novo)")

    # 5. Identifica casais
    df_todos = pd.DataFrame(continuantes + novatos_raw)
    partner  = identificar_casais(df_todos)
    print(f"   💑 Casais: {len(partner)//2}")

    # 6. Distribui novatos nas vagas abertas
    if novatos_raw:
        print(f"\n⚖️  Distribuindo novatos...")
        ocupacao = {}
        for r in continuantes:
            t = r['TRIBO']; s = r['Sexo'][0].upper()
            ocupacao.setdefault(t, {'M':0,'F':0})[s] += 1

        # Dados numéricos para score
        for r in novatos_raw:
            r['Peso']  = pd.to_numeric(str(r.get('Peso','') or ''), errors='coerce') or 0
            r['Idade'] = pd.to_numeric(str(r.get('Idade','') or ''), errors='coerce') or 0

        partner_fc = partner.copy()
        novatos_raw = distribuir_novatos(novatos_raw, ocupacao, partner_fc)
        for r in novatos_raw:
            print(f"   FC{int(r['FC']):3d} {r['Nome'].strip()} → {r['TRIBO']}")

    # 7. DataFrame final
    df_final = pd.DataFrame(continuantes + novatos_raw)
    df_final['FC'] = df_final['FC'].astype(int)

    # 8. Testes
    print(f"\n✅ Rodando testes...")
    erros, avisos = rodar_testes(df_final, partner)
    if erros:
        print("  ❌ ERROS:")
        for e in erros: print(f"     {e}")
    else:
        print("  ✅ Todos os testes passaram!")
    for a in avisos:
        print(f"  ⚠️  {a}")

    # 9. Atualiza planilha
    print(f"\n📝 Atualizando planilha...")
    ws_cad = wb['Cadastro Geral Campistas']
    ws_fp  = wb['Relatorio Familia Provas']
    ws_on  = wb['Relatorio Chamada Onibus']

    atualizar_cadastro(ws_cad, df_final, partner)
    atualizar_familia(ws_fp, df_final, partner)
    atualizar_onibus(ws_on, df_final)

    wb.save(output_path)
    print(f"\n{'='*55}")
    print(f"  ✅ PRONTO: {os.path.basename(output_path)}")
    print(f"{'='*55}")
    print(f"\n  Tribos:")
    for t in TRIBOS:
        g = df_final[df_final['TRIBO']==t]
        m = (g['Sexo']=='M').sum(); f = (g['Sexo']=='F').sum()
        print(f"    {t:10s}: {f}F {m}M")
    print()
    return {'output_path': output_path, 'erros': erros, 'avisos': avisos, 'n_campistas': len(df_final)}

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Uso: python3 atualizar_tribos.py <inscricoes.csv> [saida.xlsx]")
        sys.exit(1)
    atualizar(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
